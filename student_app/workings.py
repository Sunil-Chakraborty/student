from django.shortcuts import render
from openpyxl import workbook,load_workbook
from openpyxl.utils import get_column_letter
from django.core.mail import EmailMessage
from django.conf import settings
from . forms import PathForm
import datetime
from datetime import datetime


def ExcelView(request):

    form1 = PathForm()
    f=request.GET.get('path_name')
    print(f)
    now =  datetime.now()
    
    dd = now.strftime("%d")
    mm = now.strftime("%m")
    yy = now.strftime("%Y")
 
    if len(dd) == 1:
        dd="0"+dd
        
    if len(mm) == 1:
        mm="0"+mm
        
    yy=yy[-2:]

    path_txt="/home/backup_"+dd+mm+yy+"_1.dump"
    path_txt="/Backup/backup_"+dd+mm+yy+"_1.dump"
    int(0 if f is None else f)
    path_txt="/Backup/backup_"+f
    print(path_txt)

    email_subject = 'Email with attachment'
    email_body = 'Please find the attachment'
    from_email = settings.EMAIL_HOST_USER
    recipient_list = ['to.sunilc59@gmail.com', 'to.sunilc@gmail.com']
    
    email = EmailMessage(email_subject, email_body, from_email, recipient_list)
    
    #email.attach_file('/path/to/file.pdf')    
    #email.attach_file('/personal/Sunil_CV.docx')
    
    x="/personal/Adhar Card.pdf"
    
    #email.attach_file(x)
    y=""
    y = request.GET.get('path_name')
    y=str(y)
    print(y)
    
    email.attach_file(path_txt)
    
    #if y is not None :
    #    email.attach_file(y)
    
    #if y != "":
    #    email.attach_file(x)
        
    email.send()
    
    wb = load_workbook('catg.xlsx')
    ws = wb.active
    qs=[]
    i=0
    for row in range(99,104):
        for col in range(2,5):
            char = get_column_letter(col)
            x=ws[char + str(row)].value            
            qs.insert(row,x)                
            
    print(qs)
    """
    qs.insert(row,x)
    cell_a1 = sheet['A1'].value
    cell_b1 = sheet['B1'].value
    data = Data(cell_a1=cell_a1, cell_b1=cell_b1)
    """
    
    
    context={
        'row_iterator':range(1,6),
        'col_iterator':range(1,4),
        'queryset':qs,
        'form1':form1,
        
       }
    
    return render(request, "xls_view.html", context)

